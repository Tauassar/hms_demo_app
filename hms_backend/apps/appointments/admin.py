from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook

# Register your models here.
from apps.appointments.models import AppointmentDay, Appointment

admin.site.register(AppointmentDay)
# admin.site.register(Appointment)


@admin.register(Appointment)
class Appointment(admin.ModelAdmin):
    list_display = (
        'id',
        'patient',
        'doctor',
        'date',
        'time',
        'description',
        'medicine',
    )
    list_filter = (
        'date',
        'time',
        'patient',
        'doctor',
    )
    search_fields = (
        'id',
    )
    fieldsets = (
        ('Основное', {
            'fields': ('id', 'patient', 'doctor', 'time', 'date')
        }),
        ('Дополнительная информация', {
            'classes': ('collapse', ),
            'fields': ('description', 'medicine')
        })
    )
    readonly_fields = ('id',)
    actions = ["export_as_xlsx"]

    def export_as_xlsx(self, request, queryset):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format('report')

        filter_kwargs = {}
        for filter, arg in request.GET.items():
            if 'range' in filter:
                filter = filter.replace('range', 'date')
                arg = '-'.join(arg.split('.')[::-1])
            filter_kwargs.update({filter: arg})

        try:
            queryset = Appointment.objects.filter(**filter_kwargs)
        except:
            pass

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Appointments'
        columns = [
            'id',
            'patient',
            'doctor',
            'date',
            'time',
            'description',
            'medicine',
        ]
        # Телефон, Фамилия, Отчество, ИИН, Email

        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for obj in queryset:
            row_num += 1
            row = [
                str(obj.id),
                str(obj.patient.username),
                str(obj.doctor.username),
                str(obj.date.date),
                str(obj.time),
                str(obj.description),
                str(obj.medicine),
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    export_as_xlsx.short_description = "Скачать отчёт XLSX"
    export_as_xlsx.acts_on_all = True
